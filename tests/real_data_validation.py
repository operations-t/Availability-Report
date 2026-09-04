#!/usr/bin/env python3
"""
Optional real-source regression, replacing the old JS real-data-
validation.mjs now that model construction lives in scripts/model.py.

Runs the actual production build pipeline (the same functions
scripts/build_data.py calls) against a local folder of the five real
xlsx files, and prints summary stats for manual spot-checking --
useful for confirming a new export from Drive still produces sane
numbers before it goes live.

Usage:
    SOURCE_DIR=/path/to/folder python tests/real_data_validation.py

Expected filenames in SOURCE_DIR match Drive's production names:
    Core, Promo, KVI.xlsx
    stock.xlsx
    DOS.xlsx
    Zone Distribution.xlsx
    Ecom.xlsx
"""

import json
import os
import sys

from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "scripts"))
from model import build_model  # noqa: E402


def read_rows(path, preferred_sheet):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = preferred_sheet if preferred_sheet in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        rows = [["" if c is None else c for c in row] for row in ws.iter_rows(values_only=True)]
        return rows, sheet_name
    finally:
        wb.close()


def main():
    source_dir = os.environ.get("SOURCE_DIR")
    if not source_dir:
        sys.exit("ERROR: set SOURCE_DIR to the folder containing the five production XLSX files.")

    files = [
        ("classification", "Core, Promo, KVI.xlsx", "C-P-K"),
        ("stock", "stock.xlsx", "Sheet1"),
        ("sales", "DOS.xlsx", "DOS"),
        ("zone", "Zone Distribution.xlsx", "Final_Zone Dis"),
    ]

    inp = {"sourceMeta": {}}
    for key, filename, sheet_pref in files:
        path = os.path.join(source_dir, filename)
        rows, sheet = read_rows(path, sheet_pref)
        inp[key] = rows
        inp["sourceMeta"][key] = {
            "name": filename, "sheet": sheet,
            "rows": len(rows) - 1, "status": "Validation source",
        }
        print(f"{key}: {len(rows) - 1} data rows")

    ecom_path = os.path.join(source_dir, "Ecom.xlsx")
    wb = load_workbook(ecom_path, read_only=True, data_only=True)
    try:
        sku_sheet = "ECOM SKU" if "ECOM SKU" in wb.sheetnames else wb.sheetnames[0]
        outlet_sheet = "ECOM OUTLET" if "ECOM OUTLET" in wb.sheetnames else wb.sheetnames[1]
        ecom_sku_rows = [["" if c is None else c for c in row] for row in wb[sku_sheet].iter_rows(values_only=True)]
        ecom_outlet_rows = [["" if c is None else c for c in row] for row in wb[outlet_sheet].iter_rows(values_only=True)]
    finally:
        wb.close()
    inp["ecomSku"] = ecom_sku_rows
    inp["ecomOutlet"] = ecom_outlet_rows
    inp["sourceMeta"]["ecom"] = {
        "name": "Ecom.xlsx", "sheets": [sku_sheet, outlet_sheet],
        "skuRows": len(ecom_sku_rows) - 1, "outletRows": len(ecom_outlet_rows),
        "status": "Validation source",
    }

    model = build_model(inp, inp["sourceMeta"])

    def band_counts(codes_and_availability, low_to_high_labels):
        # Reproduces engine.js's outletAvailabilityBand thresholds for a
        # quick summary without needing the full Engine class.
        counts = {label: 0 for label in low_to_high_labels}
        for availability in codes_and_availability:
            rounded = max(0, min(100, round(availability)))
            if rounded >= 91:
                counts["91%-100%"] += 1
            elif rounded >= 81:
                counts["81%-90%"] += 1
            elif rounded >= 71:
                counts["71%-80%"] += 1
            elif rounded >= 61:
                counts["61%-70%"] += 1
            else:
                counts["Below 60%"] += 1
        return counts

    labels = ["91%-100%", "81%-90%", "71%-80%", "61%-70%", "Below 60%"]

    def summarize(stock, sales, stock_present, sales_present, required_dos=2):
        total = 0
        available = 0
        for i in range(len(stock)):
            s = stock[i]
            sa = sales[i]
            s = 0 if s != s else s
            sa = 0 if sa != sa else sa
            ads = sa / 30
            dos = None if ads == 0 else s / ads
            total += 1
            if dos is not None and dos >= required_dos:
                available += 1
        availability = (available / total * 100) if total else 0
        return {"total": total, "available": available, "unavailable": total - available,
                "availability": round(availability, 4)}

    main_summary = summarize(model["stock"], model["sales"], model["stockPresent"], model["salesPresent"])

    result = {
        "main": {
            "outlets": model["outletCount"], "skus": model["skuCount"], "slots": model["slotCount"],
            **main_summary,
        }
    }

    if model["ecom"]:
        e = model["ecom"]
        ecom_summary = summarize(e["stock"], e["sales"], e["stockPresent"], e["salesPresent"])
        result["ecom"] = {
            "listedSkus": e["listedSkuCount"], "coveredSkus": e["skuCount"],
            "dataNotCovered": e["health"]["uncoveredSkus"], "outlets": e["outletCount"],
            "slots": e["slotCount"], **ecom_summary,
        }

    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
