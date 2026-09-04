#!/usr/bin/env python3
"""
Reads the five source workbooks (downloaded by download_drive_data.py into
data/downloads/), parses them with openpyxl, runs the Python port of
model.js (build_model), and writes processed/dashboard_data.json.

Sheet selection mirrors data-loader.js's chooseSheet(): use the
configured preferred sheet name if present in the workbook, otherwise
fall back to the first sheet.
"""

import json
import os
import sys

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from model import build_model  # noqa: E402

CONFIG_PATH = "data/config.json"
DOWNLOAD_DIR = "data/downloads"
OUTPUT_PATH = "public/processed/dashboard_data.json"

DEFAULT_SOURCE_SHEETS = {
    "classification": "C-P-K",
    "sales": "DOS",
    "stock": "Sheet1",
    "zone": "Final_Zone Dis",
    "ecomSku": "ECOM SKU",
    "ecomOutlet": "ECOM OUTLET",
}

DEFAULT_FILES = {
    "classification": "core-promo-kvi.xlsx",
    "stock": "stock.xlsx",
    "sales": "dos.xlsx",
    "zone": "zone-distribution.xlsx",
    "ecom": "ecom.xlsx",
    "kviOutlet": "kvi-outlet.xlsx",
}


def load_config():
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def choose_sheet(wb, preferred):
    if preferred in wb.sheetnames:
        return preferred
    return wb.sheetnames[0]


def sheet_to_rows(ws):
    """Returns a list of lists, blank cells as ''. Mirrors xlsx-lite.js's
    rows() output shape (raw JS values: numbers stay numbers, blanks are
    empty string, everything else stringified as-is for model.py's clean()
    to normalize)."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else c for c in row])
    return rows


def read_workbook_rows(path, preferred_sheet):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = choose_sheet(wb, preferred_sheet)
        ws = wb[sheet_name]
        rows = sheet_to_rows(ws)
        return rows, sheet_name
    finally:
        wb.close()


def read_ecom_workbook(path, sku_sheet_pref, outlet_sheet_pref):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sku_sheet = choose_sheet(wb, sku_sheet_pref)
        outlet_sheet = outlet_sheet_pref if outlet_sheet_pref in wb.sheetnames else (
            wb.sheetnames[1] if len(wb.sheetnames) > 1 else None
        )
        if not outlet_sheet:
            raise ValueError(f"Ecom workbook must contain {outlet_sheet_pref}.")
        sku_rows = sheet_to_rows(wb[sku_sheet])
        outlet_rows = sheet_to_rows(wb[outlet_sheet])
        return sku_rows, sku_sheet, outlet_rows, outlet_sheet
    finally:
        wb.close()


def main():
    config = load_config()
    source_sheets = {**DEFAULT_SOURCE_SHEETS, **config.get("sourceSheets", {})}
    file_map = {**DEFAULT_FILES, **config.get("downloadedFiles", {})}

    print("Reading source workbooks...")
    inp = {"sourceMeta": {}}

    for source in ["classification", "zone", "stock", "sales"]:
        rel_path = file_map[source]
        path = os.path.join(DOWNLOAD_DIR, rel_path)
        if not os.path.exists(path):
            sys.exit(f"ERROR: expected downloaded file not found: {path}")
        rows, sheet = read_workbook_rows(path, source_sheets[source])
        inp[source] = rows
        inp["sourceMeta"][source] = {
            "name": rel_path, "sheet": sheet,
            "rows": max(len(rows) - 1, 0), "status": "Loaded (Google Drive, build-time)",
        }
        print(f"  {source}: {rel_path} -> sheet '{sheet}', {len(rows)} rows")

    ecom_path = os.path.join(DOWNLOAD_DIR, file_map["ecom"])
    if not os.path.exists(ecom_path):
        sys.exit(f"ERROR: expected downloaded file not found: {ecom_path}")
    ecom_sku_rows, sku_sheet, ecom_outlet_rows, outlet_sheet = read_ecom_workbook(
        ecom_path, source_sheets["ecomSku"], source_sheets["ecomOutlet"]
    )
    inp["ecomSku"] = ecom_sku_rows
    inp["ecomOutlet"] = ecom_outlet_rows
    inp["sourceMeta"]["ecom"] = {
        "name": file_map["ecom"], "sheets": [sku_sheet, outlet_sheet],
        "skuRows": max(len(ecom_sku_rows) - 1, 0), "outletRows": len(ecom_outlet_rows),
        "status": "Loaded (Google Drive, build-time)",
    }
    print(f"  ecom: {file_map['ecom']} -> sheets '{sku_sheet}' + '{outlet_sheet}', "
          f"{len(ecom_sku_rows)} sku rows, {len(ecom_outlet_rows)} outlet rows")

    kvi_outlet_path = os.path.join(DOWNLOAD_DIR, file_map["kviOutlet"])
    if os.path.exists(kvi_outlet_path):
        kvi_outlet_rows, kvi_sheet = read_workbook_rows(kvi_outlet_path, "Sheet1")
        inp["kviOutlet"] = kvi_outlet_rows
        inp["sourceMeta"]["kviOutlet"] = {
            "name": file_map["kviOutlet"], "sheet": kvi_sheet,
            "rows": max(len(kvi_outlet_rows) - 1, 0), "status": "Loaded (Google Drive, build-time)",
        }
        print(f"  kviOutlet: {file_map['kviOutlet']} -> sheet '{kvi_sheet}', {len(kvi_outlet_rows)} rows")
    else:
        # KVI outlet scoping is additive (change log Section 8-10): if the
        # file isn't present yet, every outlet.kvi simply defaults to
        # False rather than failing the whole build.
        inp["kviOutlet"] = []
        inp["sourceMeta"]["kviOutlet"] = {
            "name": file_map["kviOutlet"], "status": "Not found — all outlets default to non-KVI",
        }
        print(f"  kviOutlet: {kvi_outlet_path} not found — all outlets will default to kvi=False")

    print("Building outlet x SKU universe...")
    model = build_model(inp, inp["sourceMeta"])

    print(f"  outlets={model['outletCount']} skus={model['skuCount']} slots={model['slotCount']}")
    kvi_info = model["health"].get("kviOutlet") or {}
    print(f"  kviOutlet: listedCodes={kvi_info.get('listedCodes', 0)} "
          f"matchedOutlets={kvi_info.get('matchedOutlets', 0)} "
          f"codesNotInZone={len(kvi_info.get('codesNotInZone', []))}")
    if model["ecom"]:
        print(f"  ecom: outlets={model['ecom']['outletCount']} skus={model['ecom']['skuCount']} "
              f"slots={model['ecom']['slotCount']}")

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    def nan_to_none(arr):
        return [None if (isinstance(v, float) and v != v) else v for v in arr]

    output = {
        "outlets": model["outlets"],
        "skus": model["skus"],
        "stock": nan_to_none(model["stock"]),
        "sales": nan_to_none(model["sales"]),
        "stockPresent": model["stockPresent"],
        "salesPresent": model["salesPresent"],
        "outletCount": model["outletCount"],
        "skuCount": model["skuCount"],
        "slotCount": model["slotCount"],
        "health": model["health"],
        "ecom": None,
    }
    if model["ecom"]:
        e = model["ecom"]
        output["ecom"] = {
            "outlets": e["outlets"], "skus": e["skus"],
            "stock": nan_to_none(e["stock"]), "sales": nan_to_none(e["sales"]),
            "stockPresent": e["stockPresent"], "salesPresent": e["salesPresent"],
            "outletCount": e["outletCount"], "skuCount": e["skuCount"],
            "listedSkuCount": e["listedSkuCount"], "slotCount": e["slotCount"],
            "uncoveredSkus": e["uncoveredSkus"], "health": e["health"],
        }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, separators=(",", ":"))

    size_kb = os.path.getsize(OUTPUT_PATH) / 1024
    print(f"Wrote {OUTPUT_PATH} ({size_kb:.1f} KB)")


if __name__ == "__main__":
    main()
